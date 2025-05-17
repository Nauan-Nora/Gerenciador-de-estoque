from kivy.uix.treeview import TreeView
import openpyxl
import customtkinter as ctk
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox

arquivo = openpyxl.load_workbook("dados.xlsx")
ativa = arquivo['Plan1']
tema_atual = "white"

app = ctk.CTk()
app.geometry("300x420")
app.title("Gerenciador de estoque")
app.iconbitmap('assets/icon.ico')

fonte_titulo = ctk.CTkFont(family="@Adobe Kaiti Std R" ,size=20, weight="bold")
fonte_escritas = ctk.CTkFont(family="Times New Roman", size=12, weight="bold")



def cadastro_produto():
    main_frame.place_forget()

    def cadastrar_produto():

        v_nome = str(nome_entry.get())
        v_codigo = int(codigo_entry.get())
        v_preco = float(preco_entry.get())
        v_quantidade = int(quantidade_entry.get())
        v_venda = float(venda_entry.get())

        ultima_linha = ativa.max_row + 1
        ativa.cell(row=ultima_linha, column=1, value=v_nome)
        ativa.cell(row=ultima_linha, column=2, value=v_codigo)
        ativa.cell(row=ultima_linha, column=3, value=v_preco)
        ativa.cell(row=ultima_linha, column=4, value=v_quantidade)
        ativa.cell(row=ultima_linha, column=5, value=v_venda)

        arquivo.save("dados.xlsx")
        limpar_campos()
        mostra_estoque()
        nome_entry.focus_set()

    def limpar_campos():
        nome_entry.delete(0, "end")
        codigo_entry.delete(0, "end")
        preco_entry.delete(0, "end")
        quantidade_entry.delete(0, "end")
        venda_entry.delete(0, "end")
     
    def mostra_estoque():
        mostra_treeview.delete(*mostra_treeview.get_children())
        try:
            cabecalhos = [cell.value for cell in ativa[1]]
            mostra_treeview["columns"] = cabecalhos
            mostra_treeview["show"] = "headings"
            for col_index, col in enumerate(cabecalhos):
                mostra_treeview.heading(col, text=col)
                mostra_treeview.column(col, width=100)
            for row in ativa.iter_rows(min_row=2, values_only=True):
                mostra_treeview.insert('', tk.END, values=row)
        except IndexError:
            print("A planilha está vazia ou a primeira linha não contém cabeçalhos.")

    def pesquisar_item():
        termo_pesquisa = pesquisa_entry.get().lower()
        for item in mostra_treeview.get_children():
            valores = mostra_treeview.item(item, 'values')
            if valores:
                for valor in valores:
                    if termo_pesquisa in str(valor).lower():
                        mostra_treeview.see(item)
                        return

    def apagar_item():
        item_selecionado = mostra_treeview.selection()
        print(f"Itens selecionados na TreeView: {item_selecionado}")
        if not item_selecionado:
            messagebox.showerror("ATENÇÂO", "Não a nenhum item selecionado")
            return 

        item_id = item_selecionado[0]
        valores = mostra_treeview.item(item_id, 'values')
        print(f"Valores do item selecionado na TreeView: {valores}")
        if valores:
            nome_apagar = str(valores[1]).lower()
            print(f"Nome a apagar (convertido para minúsculo): {nome_apagar}")

            linhas_para_remover = []
            for row_index, row in enumerate(ativa.iter_rows(min_row=2), start=2):
                try:
                    nome_excel = str(row[1].value).lower() if row[1].value else ""
                    print(f"Nome na linha {row_index} da planilha (convertido para minúsculo): {nome_excel}")
                    if nome_apagar == nome_excel:
                        linhas_para_remover.append(row_index)
                except IndexError:
                    print(f"Erro ao acessar a coluna 'Nome' na linha {row_index} da planilha.")

            print(f"Linhas a serem removidas da planilha: {linhas_para_remover}")
            for linha in reversed(linhas_para_remover):
                try:
                    ativa.delete_rows(linha)
                except IndexError:
                    print(f"Erro ao tentar deletar a linha {linha} da planilha.")

            arquivo.save("dados.xlsx")
            mostra_estoque()
        else:
            messagebox.showerror("ERRO", "Não foi possivel obter os valores da tabela. Verifique se o dados.xlsx esta salvo no diretorio correto")
        arquivo.save("dados.xlsx")
        mostra_estoque()

    def verifica_preenchimento():
        v_nome = nome_entry.get()
        v_codigo = codigo_entry.get()
        v_preco = preco_entry.get()
        v_quantidade = quantidade_entry.get()
        v_venda = venda_entry.get()

        try:
            if v_nome == "":
                messagebox.showerror("ERRO DE PREENCIMENTO", "O campo de nome esta vazio")
            elif v_codigo == "":
                messagebox.showerror("ERRO DE PREENCIMENTO", "O compo de código esta vazio")
            elif v_preco == "":
                messagebox.showerror("ERRO DE PREENCIMENTO", " O campo de preço esta vazio")
            elif v_quantidade == "":
                messagebox.showerror("ERRO DE PREENCIMENTO", "O campo de quantidade esta vazio")
            elif v_venda == "":
                messagebox.showerror("ERRO DE PREENCIMENTO", "O campo de valor de venda esta vazio")
                pass
            else:
                cadastrar_produto()

        except:
            messagebox.showerror("ERRO", " Houve um erra não esperado, entre em contato com o suporte técnico")
        
    def menu_principal():
        frame_cadastro.grid_forget()
        frame_opcoes.grid_forget()
        area_visualizacao.grid_forget()
        frame_superior.grid_forget()
        main_frame.place(relx=0.5, rely=0.5, anchor=tk.CENTER)
        app.geometry("300x400")
        app.title("Gerenciador de estoque")
         

    ctk.set_appearance_mode(tema_atual)
    ctk.set_default_color_theme("green")

    app.geometry("830x710")
    app.title("Sistema de produtos")

    app.grid_rowconfigure(0, weight=0)
    app.grid_rowconfigure(1, weight=0) 
    app.grid_rowconfigure(2, weight=1) 
    app.grid_columnconfigure(0, weight=1)

    style = ttk.Style(app)
    style.theme_use("default")
    style.configure("Light.Treeview", background="white", foreground="black")
    style.configure("Light.Treeview.Heading", background="#f0f0f0", foreground="black")
    style.map("Light.Treeview", background=[("selected", "#aed1fc")])

    style.configure("Dark.Treeview", background="#333333", foreground="white")
    style.configure("Dark.Treeview.Heading", background="#555555", foreground="white")
    style.map("Dark.Treeview", background=[("selected", "#5699bc")])

    frame_superior = ctk.CTkFrame(app, fg_color="transparent")
    frame_superior.grid(row=0, column=0, padx=20, pady=20, sticky="ew")
    frame_superior.grid_columnconfigure(0, weight=1)
    frame_superior.grid_columnconfigure(1, weight=0)

    frame_cadastro = ctk.CTkFrame(frame_superior)
    frame_cadastro.grid(row=0, column=0, padx=(0, 10), pady=10, sticky="nsew")
    frame_cadastro.grid_columnconfigure(0, weight=1)

    titulo = ctk.CTkLabel(frame_cadastro, text="Formulario de cadastro", font=fonte_titulo)
    titulo.grid(row=0, column=0, pady=5, sticky="ew")

    nome_entry = ctk.CTkEntry(frame_cadastro, placeholder_text="Nome", font=fonte_escritas)
    nome_entry.grid(row=1, column=0, pady=5, sticky="ew")

    codigo_entry = ctk.CTkEntry(frame_cadastro, placeholder_text="Código", font=fonte_escritas)
    codigo_entry.grid(row=2, column=0, pady=5, sticky="ew")

    preco_entry = ctk.CTkEntry(frame_cadastro, placeholder_text="Custo (usar ponto ao envez de virgula)", font=fonte_escritas)
    preco_entry.grid(row=3, column=0, pady=5, sticky="ew")

    quantidade_entry = ctk.CTkEntry(frame_cadastro, placeholder_text="Quantidade", font=fonte_escritas)
    quantidade_entry.grid(row=4, column=0, pady=5, sticky="ew")

    venda_entry = ctk.CTkEntry(frame_cadastro, placeholder_text="Valor de venda (usar ponto ao envez de virgula)", font=fonte_escritas)
    venda_entry.grid(row=5, column=0, pady=5, sticky="ew")

    buton_cadastrar = ctk.CTkButton(frame_cadastro, text="Cadastrar", command=verifica_preenchimento, font=fonte_escritas)
    buton_cadastrar.grid(row=6, column=0, pady=5, sticky="ew")

    buton_limpar = ctk.CTkButton(frame_cadastro, text="Limpar formulario", command=limpar_campos, font=fonte_escritas)
    buton_limpar.grid(row=7, column=0, pady=5, sticky="ew")

    frame_opcoes = ctk.CTkFrame(frame_superior)
    frame_opcoes.grid(row=0, column=1, padx=(10, 0), pady=10, sticky="nsew")
    frame_opcoes.grid_columnconfigure(0, weight=1)
    frame_opcoes.grid_rowconfigure(1, weight=0)

    funcionalidades_label = ctk.CTkLabel(frame_opcoes, text="Funcionalidades", font=fonte_titulo)
    funcionalidades_label.grid(row=0, column=0, pady=(5, 10), padx=10, sticky="ew")

    pesquisa_entry = ctk.CTkEntry(frame_opcoes, placeholder_text="Insira um nome", font=fonte_escritas)
    pesquisa_entry.grid(row=1, column=0, pady=5, padx=10, sticky="ew")

    pesquisa_button = ctk.CTkButton(frame_opcoes, text="Pesquisar", command=pesquisar_item, font=fonte_escritas)
    pesquisa_button.grid(row=2, column=0, pady=5, padx=10, sticky="ew")

    apaga_button = ctk.CTkButton(frame_opcoes, text="Apagar Item Selecionado", command=apagar_item, font=fonte_escritas)
    apaga_button.grid(row=3, column=0, pady=(5, 10), padx=10, sticky="ew")

    volta_button = ctk.CTkButton(frame_opcoes, text="Menu Principal", command=menu_principal, font=fonte_escritas)
    volta_button.grid(row=4, column=0, pady=(20, 20), padx=20, sticky="ew")

    area_visualizacao = ctk.CTkFrame(app)
    area_visualizacao.grid(row=2, column=0, padx=20, pady=20, sticky="nsew")
    area_visualizacao.grid_columnconfigure(0, weight=1)
    area_visualizacao.grid_rowconfigure(0, weight=1)

    mostra_treeview = ttk.Treeview(
        area_visualizacao,
        columns=("Código", "Nome", "Custo", "Quantidade", "Valor de venda"),
        show="headings",
        style=f"{'Dark' if tema_atual == 'dark' else 'Light'}.Treeview"
    )
    tabelaScorll = ttk.Scrollbar(area_visualizacao, orient="vertical", command=mostra_treeview.yview)
    mostra_treeview.configure(yscrollcommand=tabelaScorll.set)
    mostra_treeview.grid(row=0, column=0, sticky="nsew")
    tabelaScorll.grid(row=0, column=1, sticky="ns")

    mostra_treeview.heading("Código", text="Código")
    mostra_treeview.heading("Nome", text="Nome")
    mostra_treeview.heading("Custo", text="Custo")
    mostra_treeview.heading("Quantidade", text="Quantidade")
    mostra_treeview.heading("Valor de venda", text="Valor de venda")
    mostra_treeview.column("#0", width=0, stretch=tk.NO)
    mostra_treeview.column("Código", anchor=tk.CENTER, width=80)
    mostra_treeview.column("Nome", anchor=tk.W, width=150)
    mostra_treeview.column("Custo", anchor=tk.E, width=80)
    mostra_treeview.column("Quantidade", anchor=tk.CENTER, width=100)
    mostra_treeview.column("Valor de venda", anchor=tk.E, width=120)

    mostra_estoque()

def mostra_estoque(treeview):
    treeview.delete(*treeview.get_children())
    try:
        cabecalhos = [cell.value for cell in ativa[1]]
        treeview["columns"] = cabecalhos
        treeview["show"] = "headings"
        for col_index, col in enumerate(cabecalhos):
            treeview.heading(col, text=col)
            treeview.column(col, width=100)
        for row in ativa.iter_rows(min_row=2, values_only=True):
            treeview.insert('', tk.END, values=row)
    except IndexError:
        print("A planilha está vazia ou a primeira linha não contém cabeçalhos.")

def janela_baixa_estoque():
    main_frame.place_forget()
    app.geometry("800x600")
    app.title("Baixa / Adição de Estoque")

    def pesquisar_item_baixa():
        termo = pesquisa_baixa_entry.get().lower()
        resultados = []
        for idx, row in enumerate(
            ativa.iter_rows(min_row=2, values_only=True),
            start=2
        ):
            if any(termo in str(v).lower() for v in row if v is not None):
                resultados.append((idx, row))
        atualizar_treeview_baixa(resultados)

    def atualizar_treeview_baixa(resultados):
        treeview_baixa.delete(*treeview_baixa.get_children())
        for row_index, row in resultados:
            treeview_baixa.insert('', tk.END, iid=str(row_index), values=row)

    def popula_full():
        full = [
            (idx, vals)
            for idx, vals in enumerate(
                ativa.iter_rows(min_row=2, values_only=True),
                start=2
            )
        ]
        atualizar_treeview_baixa(full)

    def registrar_alteracao():
        try:
            sel = treeview_baixa.selection()
            if not sel:
                messagebox.showerror("ATENÇÃO", "Selecione um item na tabela.")
                return

            row_index = int(sel[0])
            nome, codigo, custo, qtd_atual, valor_venda = treeview_baixa.item(sel[0], 'values')
            tipo = tipo_combobox.get()
            qtd_alt = int(quantidade_alteracao_entry.get())

            nova_qtd = int(qtd_atual) + int(-qtd_alt if tipo=="Baixa" else qtd_alt)
            if nova_qtd < 0:
                messagebox.showerror("ATENÇÃO", "Estoque não pode ser negativo.")
                return

            ativa.cell(row=row_index, column=4, value=nova_qtd)
            arquivo.save("dados.xlsx")
            messagebox.showinfo("SUCESSO", f"Novo estoque: {nova_qtd}")
            popula_full()

        except ValueError:
            messagebox.showerror("ERRO", "Insira valores numéricos válidos.")
        except Exception as e:
            messagebox.showerror("ERRO", f"Ocorreu um erro: {e}")

    def menu_principal():
        for w in (pesquisa_baixa_label, pesquisa_baixa_entry,
                  pesquisa_baixa_button, treeview_baixa,
                  tipo_label, tipo_combobox,
                  quantidade_alteracao_label, quantidade_alteracao_entry,
                  registrar_button, voltar_button):
            w.pack_forget()
        main_frame.place(relx=0.5, rely=0.5, anchor=tk.CENTER)
        app.geometry("300x400")
        app.title("Gerenciador de estoque")

    # --- criação dos widgets ---
    pesquisa_baixa_label = ctk.CTkLabel(app, text="Pesquisar Item:", font=fonte_titulo)
    pesquisa_baixa_label.pack(pady=5)
    pesquisa_baixa_entry = ctk.CTkEntry(
        app, placeholder_text="Nome ou Código",
        width=250, font=fonte_escritas
    )
    pesquisa_baixa_entry.pack(pady=5, padx=10)
    pesquisa_baixa_button = ctk.CTkButton(
        app, text="Pesquisar",
        command=pesquisar_item_baixa,
        font=fonte_escritas
    )
    pesquisa_baixa_button.pack(pady=5)

    treeview_baixa = ttk.Treeview(
        app,
        columns=("Nome","Código","Custo","Quantidade","Valor de venda"),
        show="headings"
    )
    for col, anchor, width in [
        ("Nome", tk.W, 150),
        ("Código", tk.CENTER,  80),
        ("Custo", tk.E,       80),
        ("Quantidade", tk.CENTER,100),
        ("Valor de venda", tk.E,120)
    ]:
        treeview_baixa.heading(col, text=col)
        treeview_baixa.column(col, anchor=anchor, width=width)
    treeview_baixa.pack(pady=10, padx=10, fill="both", expand=True)

    tipo_label = ctk.CTkLabel(app, text="Tipo de Alteração:", font=fonte_escritas)
    tipo_label.pack(pady=5)
    tipo_combobox = ctk.CTkComboBox(app, values=["Baixa","Acréscimo"])
    tipo_combobox.pack(pady=5)

    quantidade_alteracao_label = ctk.CTkLabel(
        app, text="Quantidade:", font=fonte_escritas
    )
    quantidade_alteracao_label.pack(pady=5)
    quantidade_alteracao_entry = ctk.CTkEntry(
        app, placeholder_text="Quantidade a alterar",
        font=fonte_escritas
    )
    quantidade_alteracao_entry.pack(pady=5, padx=10)

    registrar_button = ctk.CTkButton(
        app, text="Registrar Alteração",
        command=registrar_alteracao,
        font=fonte_escritas
    )
    registrar_button.pack(pady=10)

    voltar_button = ctk.CTkButton(
        app, text="Menu Principal",
        command=menu_principal,
        font=fonte_escritas
    )
    voltar_button.pack(pady=10)

    # popula imediatamente
    popula_full()

editor = None 

editor = None  

def edicao_produto():
    global editor
    main_frame.place_forget()

    app.geometry("830x710")
    app.title("Tabela editavel")

    tabela_frame = ctk.CTkFrame(app)
    tabela_frame.pack(pady=20, padx=20, fill="both", expand=True)
    tabela_frame.grid_columnconfigure(0, weight=1)
    tabela_frame.grid_rowconfigure(0, weight=1)

    treeview_editavel = ttk.Treeview(
        tabela_frame,
        columns=("Nome", "Código", "Custo", "Quantidade", "Valor de venda"),
        show="headings",
        style=f"{'Dark' if tema_atual == 'dark' else 'Light'}.Treeview"
    )
    treeview_editavel.grid(row=0, column=0, sticky="nsew")

    tabela_scroll_y = ttk.Scrollbar(tabela_frame, orient="vertical", command=treeview_editavel.yview)
    treeview_editavel.configure(yscrollcommand=tabela_scroll_y.set)
    tabela_scroll_y.grid(row=0, column=1, sticky="ns")

    tabela_scroll_x = ttk.Scrollbar(tabela_frame, orient="horizontal", command=treeview_editavel.xview)
    treeview_editavel.configure(xscrollcommand=tabela_scroll_x.set)
    tabela_scroll_x.grid(row=1, column=0, sticky="ew")

    treeview_editavel.heading("Código", text="Código")
    treeview_editavel.heading("Nome", text="Nome")
    treeview_editavel.heading("Custo", text="Custo")
    treeview_editavel.heading("Quantidade", text="Quantidade")
    treeview_editavel.heading("Valor de venda", text="Valor de venda")

    treeview_editavel.column("#0", width=0, stretch=tk.NO)
    treeview_editavel.column("Nome", anchor=tk.CENTER, width=80)
    treeview_editavel.column("Código", anchor=tk.W, width=150)
    treeview_editavel.column("Custo", anchor=tk.E, width=80)
    treeview_editavel.column("Quantidade", anchor=tk.CENTER, width=100)
    treeview_editavel.column("Valor de venda", anchor=tk.E, width=120)

    def popular_tabela_edicao():
        treeview_editavel.delete(*treeview_editavel.get_children())
        try:
            cabecalhos = [cell.value for cell in ativa[1]]
            if not treeview_editavel["columns"]:
                treeview_editavel["columns"] = cabecalhos
                treeview_editavel["show"] = "headings"
                for col_index, col in enumerate(cabecalhos):
                    treeview_editavel.heading(col, text=col)
                    treeview_editavel.column(col, width=100)
            for row in ativa.iter_rows(min_row=2, values_only=True):
                treeview_editavel.insert('', tk.END, values=row)
        except IndexError:
            print("A planilha está vazia ou a primeira linha não contém cabeçalhos.")

    editor = None

    def on_cell_click(event):
        global editor
        if editor is not None:
            editor.destroy()

        item_id = treeview_editavel.identify_row(event.y)
        column_id = treeview_editavel.identify_column(event.x)

        if item_id and column_id != '#0':
            column_index = int(column_id[1:]) - 1
            bbox = treeview_editavel.bbox(item_id, column_id)
            cell_value = treeview_editavel.item(item_id, 'values')[column_index]

            editor = ctk.CTkEntry(treeview_editavel, width=bbox[2], height=bbox[3], font=fonte_escritas) 
            editor.insert(0, cell_value)
            editor.place(x=bbox[0], y=bbox[1])
            editor.select_range(0, tk.END)
            editor.focus_set()

            old_value = cell_value

            def save_edit(event):
                new_value = editor.get()
                treeview_editavel.set(item_id, column_id, new_value)
                editor.destroy()
                if new_value != old_value:
                    salvar_edicoes()

            editor.bind("<FocusOut>", save_edit)
            editor.bind("<Return>", save_edit)

    def salvar_edicoes():
        try:
            for i, item_id in enumerate(treeview_editavel.get_children()):
                values = treeview_editavel.item(item_id, 'values')
                linha_excel = i + 2 
                for j, value in enumerate(values, start=1):
                    ativa.cell(row=linha_excel, column=j, value=value)
            arquivo.save("dados.xlsx")
            messagebox.showinfo("Sucesso", "As edições foram salvas no arquivo.")
        except Exception as e:
            messagebox.showerror("Erro ao salvar", f"Ocorreu um erro ao salvar as edições: {e}")

    def popular_tabela_edicao():
        treeview_editavel.delete(*treeview_editavel.get_children())
        try:
            cabecalhos = [cell.value for cell in ativa[1]]
            if not treeview_editavel["columns"]:
                treeview_editavel["columns"] = cabecalhos
                treeview_editavel["show"] = "headings"
                for col_index, col in enumerate(cabecalhos):
                    treeview_editavel.heading(col, text=col)
                    treeview_editavel.column(col, width=100)
            for row in ativa.iter_rows(min_row=2, values_only=True):
                treeview_editavel.insert('', tk.END, values=row)
        except IndexError:
            print("A planilha está vazia ou a primeira linha não contém cabeçalhos.")

    popular_tabela_edicao()
    treeview_editavel.bind("<Double-1>", on_cell_click)

    botao_salvar = ctk.CTkButton(tabela_frame, text="Salvar Edições", command=salvar_edicoes, font=fonte_escritas)
    botao_salvar.grid(row=2, column=0, pady=10, sticky="ew")

    def menu_principal():
        global editor
        if editor is not None:
            editor.destroy()
        tabela_frame.pack_forget()
        main_frame.place(relx=0.5, rely=0.5, anchor=tk.CENTER)
        app.geometry("300x400")
        app.title("Gerenciador de estoque")

    botao_voltar = ctk.CTkButton(tabela_frame, text="Voltar ao Menu", command=menu_principal, font=fonte_escritas)
    botao_voltar.grid(row=3, column=0, pady=10, sticky="ew")



main_frame = ctk.CTkFrame(app, fg_color="transparent")
main_frame.place(relx=0.5, rely=0.5, anchor=tk.CENTER)

label_menu = ctk.CTkLabel(main_frame, text="Menu Principal", font=fonte_titulo)
label_menu.pack(pady=10, anchor=tk.CENTER)

botao_cadas = ctk.CTkButton(main_frame, text="Sistema cadastro", width=200, height=100, command=cadastro_produto, font=fonte_titulo)
botao_cadas.pack( pady=10, anchor="se")

botao_baixa = ctk.CTkButton(main_frame, text=" Baixa/Acrecimo ", width=200, height=100, command=janela_baixa_estoque, font=fonte_titulo)
botao_baixa.pack( pady=10, anchor="se")

botao_edicao = ctk.CTkButton(main_frame, text="Edição de produtos", width=200, height=100, font=fonte_titulo, command=edicao_produto)
botao_edicao.pack(pady=10, anchor="se")

app.mainloop()